"""
애니메이션 매핑 3자 검증 결과를 엑셀로 출력
Usage: python export_anim_report.py [task_id ...]
       python export_anim_report.py --all-3rd   (3차 전체)
"""
import sys, os, json, csv, glob, struct
from pathlib import Path
from collections import defaultdict

# --- Config ---
DB_HOST = "192.168.11.201"
DB_PORT = "5432"
DB_NAME = "FA_50_KAI"
DB_USER = "kai_readonly"
DB_PASS = "kai_readonly"

VCBT_BASE = Path("E:/KAI_VCBT/fa50visualdev_new/Content/01_Visual/02_Animation/VCBT")
CSV_PATH = Path("E:/UECsvDataTableConverter/ANIMToSeq/Generated/DT_AnimIdMarkerMapping.csv")
FOLDER_JSON = Path("E:/UECsvDataTableConverter/ANIMToSeq/vcbt_folder_mapping.json")
OUTPUT_DIR = Path("E:/KAI_VCBT/fa50visualdev_new/.claude/docs")


def db_query(sql):
    """psql로 쿼리 실행, 결과를 리스트로 반환"""
    import subprocess
    env = os.environ.copy()
    env["PGPASSWORD"] = DB_PASS
    cmd = ["psql", "-h", DB_HOST, "-p", DB_PORT, "-U", DB_USER, "-d", DB_NAME,
           "-t", "-A", "-F", "|", "-c", sql]
    result = subprocess.run(cmd, capture_output=True, text=True, env=env, encoding='utf-8')
    rows = []
    for line in result.stdout.strip().split('\n'):
        if line.strip():
            rows.append(line.split('|'))
    return rows


def get_task_animations(task_id):
    """DB에서 step별 animation 조회"""
    sql = f"""
    SELECT DISTINCT ON (s.step_no) s.step_no, s.animation,
           COALESCE(m.ms_step_description, '')
    FROM tbl_step s
    LEFT JOIN tbl_main_step m ON m.ms_task_id = s.step_task_id
        AND m.ms_step_no = s.step_no AND m.ms_step_sub_no = 1
    WHERE s.step_task_id = {task_id}
    ORDER BY s.step_no, s.step_sub_no;
    """
    rows = db_query(sql)
    result = []
    for r in rows:
        if len(r) >= 3:
            try:
                anim = int(r[1]) if r[1].strip() else -1
            except ValueError:
                anim = -1
            result.append({
                'step_no': int(r[0]),
                'animation': anim,
                'description': r[2]
            })
    return result


def get_csv_mappings(task_id):
    """CSV에서 해당 task의 매핑 조회"""
    mappings = []
    if not CSV_PATH.exists():
        return mappings
    with open(CSV_PATH, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        for row in reader:
            if int(row['TaskId']) == task_id:
                mappings.append({
                    'name': row['Name'],
                    'anim_id': int(row['AnimId']),
                    'marker': row['Marker'],
                    'master_seq': row['MasterSequence']
                })
    return mappings


def get_vcbt_folder(task_id):
    """DB에서 task_tm_no 조회 후 폴더 매핑 조회"""
    if not FOLDER_JSON.exists():
        return None, None
    with open(FOLDER_JSON, 'r', encoding='utf-8') as f:
        data = json.load(f)
    mappings = data.get('mappings', data)

    # DB에서 task_tm_no 직접 조회
    rows = db_query(f"SELECT task_tm_no FROM tbl_task WHERE task_id = {task_id};")
    if rows and rows[0]:
        proc = rows[0][0].strip()
        return mappings.get(proc), proc

    return None, None


def extract_markers_from_uasset(uasset_path):
    """uasset 바이너리에서 마커(A-Z, AA-AZ 등) 추출"""
    if not os.path.exists(uasset_path):
        return []
    with open(uasset_path, 'rb') as f:
        data = f.read()

    markers = set()
    for i in range(len(data) - 10):
        # Single letter marker: length=2, char, null
        if data[i] == 2 and data[i+1] == 0 and data[i+2] == 0 and data[i+3] == 0:
            ch = data[i+4]
            if 65 <= ch <= 90 and data[i+5] == 0:
                markers.add(chr(ch))
        # Double letter marker (AA, AB...): length=3, char, char, null
        if data[i] == 3 and data[i+1] == 0 and data[i+2] == 0 and data[i+3] == 0:
            ch1 = data[i+4]
            ch2 = data[i+5]
            if 65 <= ch1 <= 90 and 65 <= ch2 <= 90 and data[i+6] == 0:
                markers.add(chr(ch1) + chr(ch2))
    return sorted(markers)


def find_master_uasset(vcbt_folder):
    """VCBT 폴더에서 *_Master.uasset 찾기"""
    if not vcbt_folder:
        return None
    folder_path = VCBT_BASE / vcbt_folder
    files = list(folder_path.glob("*_Master.uasset"))
    return str(files[0]) if files else None


def analyze_task(task_id):
    """단일 Task 3자 검증"""
    steps = get_task_animations(task_id)
    csv_maps = get_csv_mappings(task_id)
    vcbt_folder, proc_no = get_vcbt_folder(task_id)
    uasset_path = find_master_uasset(vcbt_folder)
    seq_markers = extract_markers_from_uasset(uasset_path) if uasset_path else []

    # DB animation > 0 개수
    anim_steps = [s for s in steps if s['animation'] > 0]
    db_count = len(anim_steps)
    csv_count = len(csv_maps)
    seq_count = len(seq_markers)  # 마커 N개 = 애니메이션 N개

    # 마커 매핑 (animation → marker)
    anim_to_marker = {}
    for cm in csv_maps:
        anim_to_marker[cm['anim_id']] = cm['marker']

    # 검증 결과: CSV에서 기대하는 마커가 시퀀서에 있는지 확인
    expected_markers = set()
    for cm in csv_maps:
        expected_markers.add(cm['marker'])
    missing = sorted(expected_markers - set(seq_markers))

    if csv_count == 0 and len(seq_markers) > 0:
        verdict = "DB animation 미등록"
    elif csv_count == 0:
        verdict = "시퀀서 없음"
    elif missing:
        verdict = f"누락: {', '.join(missing)}"
    elif db_count != csv_count:
        verdict = f"DB({db_count})!=CSV({csv_count})"
    else:
        verdict = "PASS"

    return {
        'task_id': task_id,
        'proc_no': proc_no,
        'vcbt_folder': vcbt_folder or 'NOT FOUND',
        'uasset': uasset_path or 'NOT FOUND',
        'total_steps': len(steps),
        'db_anim_count': db_count,
        'csv_count': csv_count,
        'seq_marker_count': len(seq_markers),
        'seq_total': seq_count,
        'verdict': verdict,
        'seq_markers': seq_markers,
        'steps': steps,
        'csv_maps': csv_maps,
        'anim_to_marker': anim_to_marker,
        'anim_steps': anim_steps,
    }


def write_excel(results, output_path):
    """검증 결과를 엑셀로 출력"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()

    # --- Sheet 1: 요약 ---
    ws_summary = wb.active
    ws_summary.title = "검증요약"

    headers = ["Task ID", "절차번호", "전체 Steps", "DB 애니메이션",
               "CSV 매핑", "시퀀서 마커", "마커 범위", "검증 결과"]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for col, h in enumerate(headers, 1):
        cell = ws_summary.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    for row_idx, r in enumerate(results, 2):
        # CSV 매핑 없거나 DB animation 미등록인 Task는 요약에서 제외
        if r['verdict'] in ('DB animation 미등록', '시퀀서 없음'):
            continue
        marker_range = f"A~{r['seq_markers'][-1]}" if r['seq_markers'] else "없음"
        values = [
            r['task_id'], r['proc_no'], r['total_steps'], r['db_anim_count'],
            r['csv_count'], r['seq_marker_count'], marker_range, r['verdict']
        ]
        for col, v in enumerate(values, 1):
            cell = ws_summary.cell(row=row_idx, column=col, value=v)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
        # 결과 컬러
        result_cell = ws_summary.cell(row=row_idx, column=8)
        if r['verdict'] == 'PASS':
            result_cell.fill = pass_fill
        else:
            result_cell.fill = fail_fill

    # 열 너비
    widths = [10, 12, 10, 12, 10, 12, 10, 40]
    for i, w in enumerate(widths, 1):
        ws_summary.column_dimensions[chr(64 + i)].width = w

    # --- Sheet 2+: 절차번호별 탭 (애니메이션 step만) ---
    warn_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    for r in results:
        # DB animation 미등록 / 시퀀서 없음은 탭 생성 안 함
        if r['verdict'] in ('DB animation 미등록', '시퀀서 없음'):
            continue
        # 시트 이름 = 절차번호 (예: 28-23-20)
        sheet_name = r['proc_no'] if r['proc_no'] else str(r['task_id'])
        # 시트 이름 31자 제한
        sheet_name = sheet_name[:31]
        ws = wb.create_sheet(sheet_name)

        detail_headers = ["Step", "Animation", "Marker", "시퀀서", "절차 내용"]
        for col, h in enumerate(detail_headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

        # 애니메이션 step만 추출 + animation 순 정렬
        anim_steps_sorted = sorted(
            [s for s in r['steps'] if s['animation'] > 0],
            key=lambda x: x['animation']
        )
        seq_marker_set = set(r['seq_markers'])

        row = 2
        for step in anim_steps_sorted:
            anim = step['animation']
            anim_id = anim // 10
            marker = r['anim_to_marker'].get(anim_id, '?')

            # 시퀀서 마커 존재 여부
            if marker in seq_marker_set:
                display_marker = marker
                seq_status = 'O'
            else:
                display_marker = marker
                seq_status = '누락'

            values = [
                step['step_no'], anim, display_marker, seq_status,
                step['description'][:100]
            ]
            for col, v in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=v)
                cell.border = thin_border
                if col <= 4:
                    cell.alignment = Alignment(horizontal='center')

            # 누락이면 빨간색
            if seq_status == '누락':
                for col in range(1, 6):
                    ws.cell(row=row, column=col).fill = fail_fill

            row += 1

        # 열 너비
        col_widths = [8, 10, 8, 8, 80]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[chr(64 + i)].width = w

    # (절차별 탭에서 열 너비 설정됨)

    # 저장
    wb.save(output_path)
    print(f"엑셀 저장: {output_path}")


def main():
    task_ids = []

    if '--all-3rd' in sys.argv:
        # 3차 언리얼 전체 17개
        task_ids = [
            334001, 334004, 328002, 329001, 329003, 321000, 328003,
            394004, 328000, 329005, 349002, 370001, 394023, 328006,
            333002, 394001, 324005
        ]
    else:
        for arg in sys.argv[1:]:
            if arg.isdigit():
                task_ids.append(int(arg))

    if not task_ids:
        print("Usage: python export_anim_report.py [task_id ...] | --all-3rd")
        return

    print(f"검증 대상: {len(task_ids)}개 Task")
    results = []
    for tid in task_ids:
        print(f"  분석 중: {tid}...", end=" ", flush=True)
        r = analyze_task(tid)
        print(r['verdict'])
        results.append(r)

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    output_path = OUTPUT_DIR / "anim_mapping_verification.xlsx"
    write_excel(results, output_path)


if __name__ == "__main__":
    main()
